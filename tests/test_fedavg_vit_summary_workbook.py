import tempfile
import unittest
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # The training-only conda environment does not include spreadsheet tools.
    Workbook = None
    load_workbook = None


@unittest.skipUnless(Workbook is not None, "openpyxl is unavailable")
class FedAvgViTSummaryWorkbookTests(unittest.TestCase):
    def test_summary_formulas_and_charts_are_written(self):
        from summarize_fedavg_vit_alpha_144 import (
            ALPHAS,
            DATASETS,
            FLOORS,
            SEEDS,
            add_floor_sheet,
            add_summary_sheet,
            add_trend_sheet,
        )

        workbook = Workbook()
        details = workbook.active
        details.title = "Run Details"
        details.append([
            "Run ID", "Dataset", "Alpha", "Min Require Size", "Partition Seed",
            "Training Seed", "Status", "Rounds", "First Round", "Final Round",
            "Best Round", "Best Acc (%)", "Best ECE", "Best NLL", "Final Acc (%)",
            "Final ECE", "Final NLL", "R96–100 Mean Acc (%)", "R96–100 SD (%)",
            "Mean Round Time (s)", "Total Round Time (s)", "Console Log", "Checkpoint",
        ])
        for dataset_index, dataset in enumerate(DATASETS):
            for floor in FLOORS:
                for alpha in ALPHAS:
                    for seed_index, seed in enumerate(SEEDS):
                        best_acc = 50.0 + dataset_index + floor + alpha + seed_index
                        details.append([
                            "fixture", dataset, alpha, floor, 42, seed, "complete",
                            100, 1, 100, 100, best_acc, 0.1, 1.0, best_acc - 0.1,
                            0.1, 1.0, best_acc - 0.2, 0.01, 2.0, 200.0,
                            "console.log", "checkpoint.pth",
                        ])

        add_summary_sheet(workbook)
        add_floor_sheet(workbook)
        add_trend_sheet(workbook)

        charts = workbook["Trend Charts"]._charts
        expected_ranges = [(2, 9), (10, 17), (18, 25)]
        for chart, (start, end) in zip(charts, expected_ranges):
            self.assertEqual(chart.series[0].val.numRef.f, f"'Trend Charts'!$C${start}:$C${end}")
            self.assertEqual(chart.series[1].val.numRef.f, f"'Trend Charts'!$D${start}:$D${end}")
            self.assertEqual(chart.series[0].cat.numRef.f, f"'Trend Charts'!$B${start}:$B${end}")

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "summary.xlsx"
            workbook.save(workbook_path)
            reopened = load_workbook(workbook_path, data_only=False)

        count_formula = reopened["3-Seed Summary"]["D2"].value
        mean_formula = reopened["3-Seed Summary"]["H2"].value
        self.assertIsInstance(count_formula, str)
        self.assertTrue(count_formula.startswith("=COUNTIFS("))
        self.assertIsInstance(mean_formula, str)
        self.assertTrue(mean_formula.startswith("="))


if __name__ == "__main__":
    unittest.main()
