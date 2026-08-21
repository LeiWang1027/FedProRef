#!/usr/bin/env python3
"""Build the audited Excel workbook for the active 171-run ViT plan."""

from __future__ import annotations

import json
import re
import statistics
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT = Path(__file__).resolve().parent
RUNNER = PROJECT / "run_vit_minpc0_methods_432.sh"
RESULTS_ROOT = PROJECT / "results" / "vit_minpc0_methods_432"
LEGACY_ROOT = PROJECT / "results"
DEFAULT_OUTPUT = Path(
    PROJECT / "results" / "vit_minpc0_methods_432" / "summary" /
    "FedProRef_client10_Tiny3alpha_171_results.xlsx"
)
CHECKPOINT_HASH = "4b8699299b1e8997753c64b052ba32031449d5d853f55a039148560ee02b820f"
ROUND_RE = re.compile(
    r"^Round\s+(\d+)\s+\|\s+acc:\s*([-+0-9.eE]+)\s+\|\s+"
    r"ece:\s*([-+0-9.eE]+)\s+\|\s+nll:\s*([-+0-9.eE]+)\s+\|\s+"
    r"round_time:\s*([-+0-9.eE]+)\s*$",
    re.MULTILINE,
)
DATASET_ORDER = {"cifar10": 0, "cifar100": 1, "tinyimagenet": 2}
METHOD_ORDER = {"fedproref": 0, "proto_aug": 1, "direct_anchor_aug": 2}
METHOD_LABEL = {
    "fedproref": "FedProRef",
    "proto_aug": "ProtoAug",
    "direct_anchor_aug": "DirectAnchorAug",
}

NAVY = "17365D"
BLUE = "5B9BD5"
LIGHT_GREEN = "E2F0D9"
LIGHT_RED = "FCE4D6"
LIGHT_BLUE = "D9EAF7"
LIGHT_GRAY = "E7E6E6"
WHITE = "FFFFFF"
BLACK = "000000"
THIN = Side(style="thin", color="B4C6E7")
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def read_plan() -> list[dict]:
    text = subprocess.check_output(
        ["bash", str(RUNNER), "plan"], cwd=PROJECT, text=True
    )
    rows = []
    for line in text.splitlines():
        fields = line.split("\t")
        if len(fields) != 8:
            raise RuntimeError(f"Unexpected plan row: {line!r}")
        run_id, dataset, alpha, total, selected, part_seed, train_seed, method = fields
        rows.append(
            {
                "run_id": run_id,
                "dataset": dataset,
                "alpha": float(alpha),
                "total_clients": int(total),
                "select_clients": int(selected),
                "partition_seed": int(part_seed),
                "training_seed": int(train_seed),
                "method": method,
            }
        )
    if len(rows) != 171:
        raise RuntimeError(f"Expected 171 planned runs, found {len(rows)}")
    return rows


def parse_rounds(path: Path) -> list[dict]:
    text = path.read_text(encoding="utf-8", errors="replace")
    return [
        {
            "round": int(match.group(1)),
            "acc": float(match.group(2)),
            "ece": float(match.group(3)),
            "nll": float(match.group(4)),
            "round_time": float(match.group(5)),
        }
        for match in ROUND_RE.finditer(text)
    ]


def current_source(row: dict):
    run_dir = RESULTS_ROOT / row["run_id"]
    config_path = run_dir / "config.json"
    log_path = run_dir / "console.log"
    if not (run_dir / ".done").is_file():
        return None
    if not config_path.is_file() or not log_path.is_file():
        return None
    try:
        config = json.loads(config_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    if config.get("checkpoint_hash") != CHECKPOINT_HASH:
        return None
    if config.get("min_require_size") != 0:
        return None
    if config.get("partition_scheme") != "partition-nonempty-v4":
        return None
    rounds = parse_rounds(log_path)
    if not any(item["round"] == 100 for item in rounds):
        return None
    return "current", log_path, config, rounds


def legacy_source(row: dict):
    if row["dataset"] not in {"cifar100", "tinyimagenet"}:
        return None
    if row["alpha"] != 0.01 or row["total_clients"] != 10:
        return None
    if row["method"] not in {"fedproref", "proto_aug"}:
        return None
    run_dir = LEGACY_ROOT / (
        f"mcstress_{row['dataset']}_a001_ps42_ts{row['training_seed']}_{row['method']}"
    )
    log_path = run_dir / "console.log"
    command_path = run_dir / "command.sh"
    if not (run_dir / ".done").is_file():
        return None
    if not log_path.is_file() or not command_path.is_file():
        return None
    command = command_path.read_text(encoding="utf-8", errors="replace")
    required = [
        "--backbone ViT-B-16",
        "old_open_clip_model.safetensors",
        "--comm_rounds 100",
        "--local_epochs 10",
        f"--method {row['method']}",
        f"--seed {row['training_seed']}",
        "--partition_seed 42",
        "--num_clients 10",
        "--select_clients 10",
    ]
    if not all(token in command for token in required):
        return None
    rounds = parse_rounds(log_path)
    if not any(item["round"] == 100 for item in rounds):
        return None
    config = {
        "backbone": "ViT-B-16",
        "local_epochs": 10,
        "comm_rounds": 100,
    }
    return "legacy", log_path, config, rounds


def build_details(plan: list[dict]) -> list[dict]:
    details = []
    for plan_row in plan:
        row = dict(plan_row)
        source = current_source(row) or legacy_source(row)
        if source is None:
            row.update(
                {
                    "status": "缺失",
                    "source": "",
                    "rounds": 0,
                    "final_acc": None,
                    "best_acc": None,
                    "best_round": None,
                    "best_ece": None,
                    "best_nll": None,
                    "r96_100_mean": None,
                    "r96_100_sd": None,
                    "mean_round_time": None,
                    "total_round_time": None,
                    "final_ece": None,
                    "final_nll": None,
                    "backbone": "ViT-B-16",
                    "local_epochs": 10,
                    "comm_rounds": 100,
                    "log_path": "",
                    "note": "当前正式矩阵中尚无严格完成日志",
                }
            )
            details.append(row)
            continue

        source_name, log_path, config, rounds = source
        if len(rounds) != 100:
            raise RuntimeError(f"Expected 100 rounds, found {len(rounds)}: {log_path}")
        round_numbers = [item["round"] for item in rounds]
        if round_numbers != list(range(1, 101)):
            raise RuntimeError(f"Rounds are not contiguous 1..100: {log_path}")
        final = rounds[-1]
        best = max(rounds, key=lambda item: item["acc"])
        last_five = [item["acc"] for item in rounds if 96 <= item["round"] <= 100]
        row.update(
            {
                "status": "完成",
                "source": source_name,
                "rounds": 100,
                "final_acc": final["acc"],
                "best_acc": best["acc"],
                "best_round": best["round"],
                "best_ece": best["ece"],
                "best_nll": best["nll"],
                "r96_100_mean": statistics.mean(last_five),
                "r96_100_sd": statistics.stdev(last_five),
                "mean_round_time": statistics.mean(item["round_time"] for item in rounds),
                "total_round_time": sum(item["round_time"] for item in rounds),
                "final_ece": final["ece"],
                "final_nll": final["nll"],
                "backbone": config.get("backbone", "ViT-B-16"),
                "local_epochs": int(config.get("local_epochs", 10)),
                "comm_rounds": int(config.get("comm_rounds", 100)),
                "log_path": str(log_path),
                "note": "协议匹配的历史正式日志" if source_name == "legacy" else "",
            }
        )
        details.append(row)

    details.sort(
        key=lambda item: (
            DATASET_ORDER[item["dataset"]],
            item["alpha"],
            METHOD_ORDER[item["method"]],
            item["training_seed"],
        )
    )
    return details


def title(sheet, text: str, end_col: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = sheet.cell(1, 1, text)
    cell.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 25


def header(sheet, row: int, end_col: int) -> None:
    for cell in sheet[row][:end_col]:
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = ALL_BORDER
    sheet.row_dimensions[row].height = 32


def add_table(sheet, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def set_default_font(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows():
            for cell in row:
                if cell.row != 1 and cell.font.name != "Arial":
                    cell.font = Font(
                        name="Arial",
                        size=10,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color,
                    )
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical or "center",
                    wrap_text=cell.alignment.wrap_text,
                )


def build_workbook(details: list[dict], output: Path) -> None:
    workbook = Workbook()
    detail_ws = workbook.active
    detail_ws.title = "实验明细"
    summary_ws = workbook.create_sheet("三Seed汇总")
    audit_ws = workbook.create_sheet("审计")
    plan_ws = workbook.create_sheet("计划与口径")
    exclude_ws = workbook.create_sheet("排除说明")

    detail_headers = [
        "序号",
        "Run ID",
        "状态",
        "日志来源",
        "数据集",
        "Alpha",
        "Total Clients",
        "Select Clients",
        "Partition Seed",
        "Training Seed",
        "方法",
        "完成轮数",
        "最终准确率(%)",
        "最佳准确率(%)",
        "最佳轮次",
        "最佳轮ECE",
        "最佳轮NLL",
        "R96-100均值(%)",
        "R96-100样本SD",
        "平均单轮时间(s)",
        "总轮时间(s)",
        "最终ECE",
        "最终NLL",
        "Backbone",
        "Local Epochs",
        "Comm Rounds",
        "Console Log",
        "备注",
    ]
    title(detail_ws, "FedProRef client=10 正式实验明细（171组）", len(detail_headers))
    for column, value in enumerate(detail_headers, 1):
        detail_ws.cell(3, column, value)
    header(detail_ws, 3, len(detail_headers))

    for index, item in enumerate(details, 1):
        row_number = index + 3
        values = [
            index,
            item["run_id"],
            item["status"],
            item["source"],
            item["dataset"],
            item["alpha"],
            item["total_clients"],
            item["select_clients"],
            item["partition_seed"],
            item["training_seed"],
            METHOD_LABEL[item["method"]],
            item["rounds"],
            item["final_acc"],
            item["best_acc"],
            item["best_round"],
            item["best_ece"],
            item["best_nll"],
            item["r96_100_mean"],
            item["r96_100_sd"],
            item["mean_round_time"],
            item["total_round_time"],
            item["final_ece"],
            item["final_nll"],
            item["backbone"],
            item["local_epochs"],
            item["comm_rounds"],
            item["log_path"],
            item["note"],
        ]
        for column, value in enumerate(values, 1):
            cell = detail_ws.cell(row_number, column, value)
            cell.border = ALL_BORDER
            cell.alignment = Alignment(
                vertical="center", wrap_text=column in {2, 27, 28}
            )
        detail_ws.cell(row_number, 3).fill = PatternFill(
            "solid", fgColor=LIGHT_GREEN if item["status"] == "完成" else LIGHT_RED
        )
        if item["log_path"]:
            log_cell = detail_ws.cell(row_number, 27)
            log_cell.hyperlink = "file://" + quote(item["log_path"])
            log_cell.font = Font(
                name="Arial", size=10, color="0563C1", underline="single"
            )
    for row_number in range(4, len(details) + 4):
        detail_ws.cell(row_number, 6).number_format = "0.00"
        for column in range(13, 24):
            detail_ws.cell(row_number, column).number_format = "0.0000"
    detail_ws["M3"].comment = Comment(
        "准确率存储为百分数值，例如93.65表示93.65%。", "Codex"
    )
    detail_ws["N3"].comment = Comment(
        "从该实验完整100轮测试集准确率中取最大值。", "Codex"
    )
    detail_ws.freeze_panes = "A4"
    detail_ws.auto_filter.ref = f"A3:AB{len(details) + 3}"
    add_table(detail_ws, f"A3:AB{len(details) + 3}", "FormalRunDetails")
    widths = [
        7, 68, 10, 11, 16, 9, 13, 13, 14, 13, 18, 10, 15, 15,
        10, 12, 12, 17, 17, 17, 15, 12, 12, 14, 12, 12, 85, 28,
    ]
    for column, width in enumerate(widths, 1):
        detail_ws.column_dimensions[get_column_letter(column)].width = width

    summary_headers = [
        "数据集",
        "Alpha",
        "方法",
        "预期Seed数",
        "完成Seed数",
        "最佳准确率均值(%)",
        "最佳准确率SD",
        "最终准确率均值(%)",
        "最终准确率SD",
        "R96-100均值(%)",
        "R96-100跨Seed SD",
        "平均单轮时间(s)",
    ]
    title(summary_ws, "三 Seed 汇总（均值与样本标准差）", len(summary_headers))
    for column, value in enumerate(summary_headers, 1):
        summary_ws.cell(3, column, value)
    header(summary_ws, 3, len(summary_headers))

    groups: dict[tuple, list[int]] = defaultdict(list)
    for detail_row, item in enumerate(details, 4):
        groups[(item["dataset"], item["alpha"], item["method"])].append(detail_row)
    summary_row = 4
    for key in sorted(
        groups, key=lambda item: (DATASET_ORDER[item[0]], item[1], METHOD_ORDER[item[2]])
    ):
        dataset, alpha, method = key
        rows = groups[key]
        if len(rows) != 3 or rows != list(range(rows[0], rows[0] + 3)):
            raise RuntimeError(f"Expected contiguous three-seed group for {key}: {rows}")
        start, end = rows[0], rows[-1]
        summary_ws.cell(summary_row, 1, dataset)
        summary_ws.cell(summary_row, 2, alpha)
        summary_ws.cell(summary_row, 3, METHOD_LABEL[method])
        summary_ws.cell(summary_row, 4, 3)
        summary_ws.cell(summary_row, 5, f"=COUNT('实验明细'!N{start}:N{end})")
        summary_ws.cell(
            summary_row, 6,
            f"=IFERROR(AVERAGE('实验明细'!N{start}:N{end}),\"\")",
        )
        summary_ws.cell(
            summary_row, 7,
            f"=IF(COUNT('实验明细'!N{start}:N{end})>1,STDEV('实验明细'!N{start}:N{end}),\"\")",
        )
        summary_ws.cell(
            summary_row, 8,
            f"=IFERROR(AVERAGE('实验明细'!M{start}:M{end}),\"\")",
        )
        summary_ws.cell(
            summary_row, 9,
            f"=IF(COUNT('实验明细'!M{start}:M{end})>1,STDEV('实验明细'!M{start}:M{end}),\"\")",
        )
        summary_ws.cell(
            summary_row, 10,
            f"=IFERROR(AVERAGE('实验明细'!R{start}:R{end}),\"\")",
        )
        summary_ws.cell(
            summary_row, 11,
            f"=IF(COUNT('实验明细'!R{start}:R{end})>1,STDEV('实验明细'!R{start}:R{end}),\"\")",
        )
        summary_ws.cell(
            summary_row, 12,
            f"=IFERROR(AVERAGE('实验明细'!T{start}:T{end}),\"\")",
        )
        for column in range(1, len(summary_headers) + 1):
            summary_ws.cell(summary_row, column).border = ALL_BORDER
        summary_row += 1
    for row_number in range(4, summary_row):
        summary_ws.cell(row_number, 2).number_format = "0.00"
        for column in range(6, 13):
            summary_ws.cell(row_number, column).number_format = "0.0000"
    summary_ws.freeze_panes = "A4"
    add_table(summary_ws, f"A3:L{summary_row - 1}", "ThreeSeedSummary")
    summary_widths = [16, 9, 18, 13, 13, 20, 18, 20, 18, 18, 20, 18]
    for column, width in enumerate(summary_widths, 1):
        summary_ws.column_dimensions[get_column_letter(column)].width = width

    audit_headers = ["检查项", "期望值", "工作簿计算值", "结论"]
    title(audit_ws, "数据完整性审计", len(audit_headers))
    for column, value in enumerate(audit_headers, 1):
        audit_ws.cell(3, column, value)
    header(audit_ws, 3, len(audit_headers))
    first_detail, last_detail = 4, len(details) + 3
    audit_items = [
        ("正式计划行数", 171, f"=COUNTA('实验明细'!B{first_detail}:B{last_detail})"),
        ("完成实验数", 171, f'=COUNTIF(\'实验明细\'!C{first_detail}:C{last_detail},"完成")'),
        ("当前严格日志", 159, f'=COUNTIF(\'实验明细\'!D{first_detail}:D{last_detail},"current")'),
        ("协议匹配历史日志", 12, f'=COUNTIF(\'实验明细\'!D{first_detail}:D{last_detail},"legacy")'),
        ("100轮完整日志", 171, f'=COUNTIF(\'实验明细\'!L{first_detail}:L{last_detail},100)'),
        ("包含最佳测试准确率", 171, f"=COUNT('实验明细'!N{first_detail}:N{last_detail})"),
        ("Total-client=50正式行", 0, f'=COUNTIF(\'实验明细\'!G{first_detail}:G{last_detail},50)'),
        ("TinyImageNet正式行", 27, f'=COUNTIF(\'实验明细\'!E{first_detail}:E{last_detail},"tinyimagenet")'),
        (
            "TinyImageNet非法Alpha行",
            0,
            f'=SUMPRODUCT((\'实验明细\'!E{first_detail}:E{last_detail}="tinyimagenet")*'
            f'(\'实验明细\'!F{first_detail}:F{last_detail}<>0.01)*'
            f'(\'实验明细\'!F{first_detail}:F{last_detail}<>0.03)*'
            f'(\'实验明细\'!F{first_detail}:F{last_detail}<>0.05))',
        ),
    ]
    for offset, (name, expected, formula) in enumerate(audit_items, 4):
        audit_ws.cell(offset, 1, name)
        audit_ws.cell(offset, 2, expected)
        audit_ws.cell(offset, 3, formula)
        audit_ws.cell(offset, 4, f'=IF(B{offset}=C{offset},"通过","检查")')
        for column in range(1, 5):
            audit_ws.cell(offset, column).border = ALL_BORDER
        audit_ws.cell(offset, 4).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    audit_ws.column_dimensions["A"].width = 30
    audit_ws.column_dimensions["B"].width = 15
    audit_ws.column_dimensions["C"].width = 18
    audit_ws.column_dimensions["D"].width = 15

    title(plan_ws, "实验计划与统计口径", 4)
    plan_ws.append([])
    plan_ws.append(["项目", "取值/说明", "来源", "备注"])
    header(plan_ws, 3, 4)
    plan_items = [
        ("正式实验总数", 171, "当前SH矩阵", "只统计当前计划"),
        ("数据集", "CIFAR-10、CIFAR-100、TinyImageNet", "当前SH矩阵", ""),
        ("CIFAR Alpha", "0.01, 0.03, 0.05, 0.07, 0.09, 0.10, 0.30, 0.50", "当前SH矩阵", "两数据集均为8值"),
        ("TinyImageNet Alpha", "0.01, 0.03, 0.05", "用户确认", "其他Alpha暂不运行"),
        ("方法", "FedProRef、ProtoAug、DirectAnchorAug", "当前SH矩阵", ""),
        ("训练Seed", "42, 43, 44", "当前SH矩阵", ""),
        ("Partition Seed", 42, "当前SH矩阵", "固定"),
        ("Total Clients", 10, "用户确认", "c50已排除"),
        ("Select Clients", 10, "当前SH矩阵", ""),
        ("通信轮数", 100, "当前SH矩阵", ""),
        ("Local Epochs", 10, "当前SH矩阵", ""),
        ("Backbone", "ViT-B/16", "当前SH矩阵", "加载本地预训练权重"),
        ("最佳准确率口径", "每个实验100轮测试集准确率的最大值", "Console Log", "同时记录对应轮次/ECE/NLL"),
        ("三Seed SD", "样本标准差（STDEV）", "三Seed汇总公式", "n=3"),
        ("生成时间", datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %Z"), "本机时间", ""),
    ]
    for row_number, values in enumerate(plan_items, 4):
        for column, value in enumerate(values, 1):
            cell = plan_ws.cell(row_number, column, value)
            cell.border = ALL_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    add_table(plan_ws, f"A3:D{len(plan_items) + 3}", "PlanDefinition")
    for column, width in enumerate([26, 72, 22, 34], 1):
        plan_ws.column_dimensions[get_column_letter(column)].width = width

    title(exclude_ws, "不纳入当前正式统计的结果", 5)
    exclude_headers = ["排除范围", "状态", "是否删除", "是否纳入171组", "说明"]
    for column, value in enumerate(exclude_headers, 1):
        exclude_ws.cell(3, column, value)
    header(exclude_ws, 3, len(exclude_headers))
    exclusions = [
        (
            "所有 total-client=50 实验",
            "暂不运行/不统计",
            "否",
            "否",
            "已有日志和结果保留在原目录，不进入本工作簿正式明细。",
        ),
        (
            "TinyImageNet alpha=0.07,0.09,0.10,0.30,0.50",
            "暂不运行/不统计",
            "否",
            "否",
            "根据用户确认，仅保留0.01、0.03、0.05。",
        ),
    ]
    for row_number, values in enumerate(exclusions, 4):
        for column, value in enumerate(values, 1):
            cell = exclude_ws.cell(row_number, column, value)
            cell.border = ALL_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    add_table(exclude_ws, "A3:E5", "ExclusionRules")
    for column, width in enumerate([52, 20, 12, 18, 72], 1):
        exclude_ws.column_dimensions[get_column_letter(column)].width = width

    set_default_font(workbook)
    workbook.properties.title = "FedProRef client=10 TinyImageNet 3-alpha results"
    workbook.properties.subject = "171-run audited experiment workbook"
    workbook.properties.creator = "Codex"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def main() -> int:
    output = Path(sys.argv[1]).expanduser().resolve() if len(sys.argv) > 1 else DEFAULT_OUTPUT
    if output.exists():
        raise RuntimeError(f"Refusing to overwrite existing workbook: {output}")
    plan = read_plan()
    details = build_details(plan)
    source_counts = Counter(item["source"] for item in details)
    status_counts = Counter(item["status"] for item in details)
    if status_counts["完成"] != 171:
        raise RuntimeError(f"Expected 171 complete runs: {dict(status_counts)}")
    if source_counts["current"] != 159 or source_counts["legacy"] != 12:
        raise RuntimeError(f"Unexpected source counts: {dict(source_counts)}")
    if len({item["run_id"] for item in details}) != 171:
        raise RuntimeError("Duplicate Run IDs detected")
    if any(item["total_clients"] != 10 for item in details):
        raise RuntimeError("A c50 run entered the formal plan")
    allowed_tiny = {0.01, 0.03, 0.05}
    if any(
        item["dataset"] == "tinyimagenet" and item["alpha"] not in allowed_tiny
        for item in details
    ):
        raise RuntimeError("An excluded TinyImageNet alpha entered the formal plan")
    build_workbook(details, output)
    print(
        json.dumps(
            {
                "output": str(output),
                "formal_runs": len(details),
                "complete": status_counts["完成"],
                "current": source_counts["current"],
                "legacy": source_counts["legacy"],
                "missing": status_counts["缺失"],
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
