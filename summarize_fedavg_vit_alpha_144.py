#!/usr/bin/env python3
"""Summarize the completed 144-run FedAvg ViT-B/16 experiment matrix."""

from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from statistics import mean, median, stdev

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
RESULTS_ROOT = ROOT / "results" / "fedavg_vit_alpha_144"
OUTPUT_ROOT = ROOT / "summary" / "fedavg_vit_alpha_144"

DATASETS = ("cifar10", "cifar100", "tinyimagenet")
ALPHAS = (0.01, 0.03, 0.05, 0.07, 0.09, 0.10, 0.30, 0.50)
FLOORS = (0, 1)
SEEDS = (42, 43, 44)

ROUND_RE = re.compile(
    r"^Round\s+(\d+)\s+\|\s+acc:\s*([-+0-9.eE]+)\s+\|\s+"
    r"ece:\s*([-+0-9.eE]+)\s+\|\s+nll:\s*([-+0-9.eE]+)\s+\|\s+"
    r"round_time:\s*([-+0-9.eE]+)$",
    re.MULTILINE,
)
BEST_RE = re.compile(
    r"^\s*Best:\s*Round\s+(\d+)\s*\|\s*Acc=([-+0-9.eE]+)%\s*\|\s*"
    r"ECE=([-+0-9.eE]+)\s*\|\s*NLL=([-+0-9.eE]+)$",
    re.MULTILINE,
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
GOOD_FILL = PatternFill("solid", fgColor="E2F0D9")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_BOLD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Arial", size=10, color="000000")
INPUT_FONT = Font(name="Arial", size=10, color="0000FF")
FORMULA_FONT = Font(name="Arial", size=10, color="000000")


def run_id(dataset: str, alpha: float, floor: int, seed: int) -> str:
    tag = f"a{int(round(alpha * 100)):03d}"
    return f"vit_fedavg_{dataset}_{tag}_minpc{floor}_ps42_ts{seed}"


def expected_matrix() -> list[tuple[str, str, float, int, int, int]]:
    return [
        (run_id(dataset, alpha, floor, seed), dataset, alpha, floor, 42, seed)
        for dataset in DATASETS
        for floor in FLOORS
        for alpha in ALPHAS
        for seed in SEEDS
    ]


def parse_run(row: tuple[str, str, float, int, int, int]) -> dict[str, object]:
    rid, dataset, alpha, floor, partition_seed, training_seed = row
    run_dir = RESULTS_ROOT / rid
    config_path = run_dir / "config.json"
    log_path = run_dir / "console.log"
    done_path = run_dir / ".done"
    checkpoint_path = run_dir / "checkpoints" / f"{rid}_fedavg.pth"

    config = json.loads(config_path.read_text(encoding="utf-8")) if config_path.is_file() else {}
    content = log_path.read_text(encoding="utf-8", errors="replace") if log_path.is_file() else ""
    rounds = [
        {
            "round": int(match.group(1)),
            "acc": float(match.group(2)),
            "ece": float(match.group(3)),
            "nll": float(match.group(4)),
            "round_time": float(match.group(5)),
        }
        for match in ROUND_RE.finditer(content)
    ]
    best_matches = list(BEST_RE.finditer(content))
    best = best_matches[-1] if best_matches else None
    round_numbers = [int(item["round"]) for item in rounds]
    last_five = [item for item in rounds if 96 <= int(item["round"]) <= 100]

    config_ok = (
        config.get("run_id") == rid
        and config.get("dataset") == dataset
        and abs(float(config.get("alpha", -1)) - alpha) < 1e-12
        and int(config.get("min_require_size", -1)) == floor
        and int(config.get("partition_seed", -1)) == partition_seed
        and int(config.get("training_seed", -1)) == training_seed
        and config.get("method") == "fedavg"
        and config.get("backbone") == "ViT-B-16"
    )
    complete = (
        done_path.is_file()
        and log_path.is_file()
        and config_ok
        and round_numbers == list(range(1, 101))
        and len(last_five) == 5
        and best is not None
        and checkpoint_path.is_file()
    )

    final = rounds[-1] if rounds else None
    max_round = max(rounds, key=lambda item: float(item["acc"])) if rounds else None
    best_round = int(best.group(1)) if best else None
    best_acc = float(best.group(2)) if best else None
    best_ece = float(best.group(3)) if best else None
    best_nll = float(best.group(4)) if best else None
    best_matches_rounds = bool(
        best and max_round
        and best_round == int(max_round["round"])
        and abs(best_acc - float(max_round["acc"])) < 5e-5
    )

    return {
        "run_id": rid,
        "dataset": dataset,
        "alpha": alpha,
        "min_require_size": floor,
        "partition_seed": partition_seed,
        "training_seed": training_seed,
        "status": "complete" if complete else "incomplete",
        "config_ok": config_ok,
        "done_marker": done_path.is_file(),
        "checkpoint_exists": checkpoint_path.is_file(),
        "round_count": len(rounds),
        "unique_round_count": len(set(round_numbers)),
        "first_round": round_numbers[0] if round_numbers else None,
        "final_round": round_numbers[-1] if round_numbers else None,
        "best_round": best_round,
        "best_acc_pct": best_acc,
        "best_ece": best_ece,
        "best_nll": best_nll,
        "best_matches_round_log": best_matches_rounds,
        "final_acc_pct": float(final["acc"]) if final else None,
        "final_ece": float(final["ece"]) if final else None,
        "final_nll": float(final["nll"]) if final else None,
        "rounds96_100_acc_mean_pct": mean(float(item["acc"]) for item in last_five) if last_five else None,
        "rounds96_100_acc_sd_pct": stdev(float(item["acc"]) for item in last_five) if len(last_five) > 1 else None,
        "mean_round_time_s": mean(float(item["round_time"]) for item in rounds) if rounds else None,
        "total_round_time_s": sum(float(item["round_time"]) for item in rounds),
        "console_log": str(log_path.relative_to(ROOT)),
        "checkpoint": str(checkpoint_path.relative_to(ROOT)),
    }


RUN_FIELDS = (
    "run_id", "dataset", "alpha", "min_require_size", "partition_seed",
    "training_seed", "status", "config_ok", "done_marker", "checkpoint_exists",
    "round_count", "unique_round_count", "first_round", "final_round",
    "best_round", "best_acc_pct", "best_ece", "best_nll",
    "best_matches_round_log", "final_acc_pct", "final_ece", "final_nll",
    "rounds96_100_acc_mean_pct", "rounds96_100_acc_sd_pct",
    "mean_round_time_s", "total_round_time_s", "console_log", "checkpoint",
)


def write_csv(path: Path, rows: list[dict[str, object]], fields: tuple[str, ...]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def aggregate_runs(runs: list[dict[str, object]]) -> list[dict[str, object]]:
    output = []
    for dataset in DATASETS:
        for floor in FLOORS:
            for alpha in ALPHAS:
                group = [
                    row for row in runs
                    if row["dataset"] == dataset
                    and row["min_require_size"] == floor
                    and abs(float(row["alpha"]) - alpha) < 1e-12
                    and row["status"] == "complete"
                ]
                group.sort(key=lambda item: int(item["training_seed"]))
                best_values = [float(row["best_acc_pct"]) for row in group]
                final_values = [float(row["final_acc_pct"]) for row in group]
                last_five_values = [float(row["rounds96_100_acc_mean_pct"]) for row in group]
                output.append({
                    "dataset": dataset,
                    "alpha": alpha,
                    "min_require_size": floor,
                    "n_complete": len(group),
                    "best_acc_seed42": best_values[0] if len(best_values) == 3 else None,
                    "best_acc_seed43": best_values[1] if len(best_values) == 3 else None,
                    "best_acc_seed44": best_values[2] if len(best_values) == 3 else None,
                    "best_acc_mean_pct": mean(best_values) if best_values else None,
                    "best_acc_sample_sd_pct": stdev(best_values) if len(best_values) > 1 else None,
                    "best_acc_min_pct": min(best_values) if best_values else None,
                    "best_acc_max_pct": max(best_values) if best_values else None,
                    "final_acc_mean_pct": mean(final_values) if final_values else None,
                    "final_acc_sample_sd_pct": stdev(final_values) if len(final_values) > 1 else None,
                    "rounds96_100_acc_mean_pct": mean(last_five_values) if last_five_values else None,
                    "mean_round_time_s": mean(float(row["mean_round_time_s"]) for row in group) if group else None,
                    "total_round_time_h": sum(float(row["total_round_time_s"]) for row in group) / 3600 if group else None,
                })
    return output


SUMMARY_FIELDS = (
    "dataset", "alpha", "min_require_size", "n_complete",
    "best_acc_seed42", "best_acc_seed43", "best_acc_seed44",
    "best_acc_mean_pct", "best_acc_sample_sd_pct", "best_acc_min_pct",
    "best_acc_max_pct", "final_acc_mean_pct", "final_acc_sample_sd_pct",
    "rounds96_100_acc_mean_pct", "mean_round_time_s", "total_round_time_h",
)


def build_floor_pairs(runs: list[dict[str, object]]) -> list[dict[str, object]]:
    lookup = {
        (str(row["dataset"]), float(row["alpha"]), int(row["min_require_size"]), int(row["training_seed"])): row
        for row in runs if row["status"] == "complete"
    }
    output = []
    for dataset in DATASETS:
        for alpha in ALPHAS:
            for seed in SEEDS:
                floor0 = lookup[(dataset, alpha, 0, seed)]
                floor1 = lookup[(dataset, alpha, 1, seed)]
                output.append({
                    "dataset": dataset,
                    "alpha": alpha,
                    "training_seed": seed,
                    "floor0_best_acc_pct": floor0["best_acc_pct"],
                    "floor1_best_acc_pct": floor1["best_acc_pct"],
                    "floor1_minus_floor0_best_pp": float(floor1["best_acc_pct"]) - float(floor0["best_acc_pct"]),
                    "floor0_rounds96_100_pct": floor0["rounds96_100_acc_mean_pct"],
                    "floor1_rounds96_100_pct": floor1["rounds96_100_acc_mean_pct"],
                    "floor1_minus_floor0_rounds96_100_pp": (
                        float(floor1["rounds96_100_acc_mean_pct"])
                        - float(floor0["rounds96_100_acc_mean_pct"])
                    ),
                })
    return output


PAIR_FIELDS = (
    "dataset", "alpha", "training_seed", "floor0_best_acc_pct",
    "floor1_best_acc_pct", "floor1_minus_floor0_best_pp",
    "floor0_rounds96_100_pct", "floor1_rounds96_100_pct",
    "floor1_minus_floor0_rounds96_100_pp",
)


def style_table(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32


def set_widths(ws, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def add_readme_sheet(wb: Workbook, runs: list[dict[str, object]]) -> None:
    ws = wb.active
    ws.title = "README"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "FedAvg ViT-B/16 Alpha Sweep — 144-Run Summary"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="1F4E78")
    rows = [
        ("Generated", datetime.now().astimezone().isoformat(timespec="seconds")),
        ("Source", "results/fedavg_vit_alpha_144/<run_id>/console.log and config.json"),
        ("Matrix", "3 datasets × 8 alpha values × 2 min_require_size values × 3 training seeds = 144 runs"),
        ("Fixed protocol", "FedAvg; frozen OpenCLIP ViT-B/16; partition seed 42; 10/10 clients; 100 rounds; 10 local epochs"),
        ("Primary metric", "Best global test accuracy (%) observed within 100 rounds"),
        ("Secondary metrics", "Round-100 accuracy and mean accuracy over rounds 96–100"),
        ("Replicate unit", "Training seed (42, 43, 44) within one dataset/alpha/floor configuration"),
        ("Uncertainty", "Sample standard deviation across the three training seeds"),
        ("Completed", f"{sum(row['status'] == 'complete' for row in runs)} / {len(runs)}"),
        ("Blue text", "Hardcoded values parsed from source logs/configuration"),
        ("Black text", "Workbook formulas derived from Run Details"),
    ]
    for index, (label, value) in enumerate(rows, start=3):
        ws.cell(index, 1, label).font = Font(name="Arial", size=10, bold=True)
        ws.cell(index, 1).fill = SUBHEADER_FILL
        ws.cell(index, 2, value).font = INPUT_FONT if label not in {"Black text"} else FORMULA_FONT
        ws.cell(index, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110


def add_run_sheet(wb: Workbook, runs: list[dict[str, object]]) -> None:
    ws = wb.create_sheet("Run Details")
    headers = [
        "Run ID", "Dataset", "Alpha", "Min Require Size", "Partition Seed",
        "Training Seed", "Status", "Rounds", "First Round", "Final Round",
        "Best Round", "Best Acc (%)", "Best ECE", "Best NLL", "Final Acc (%)",
        "Final ECE", "Final NLL", "R96–100 Mean Acc (%)", "R96–100 SD (%)",
        "Mean Round Time (s)", "Total Round Time (s)", "Console Log", "Checkpoint",
    ]
    ws.append(headers)
    for row in runs:
        ws.append([
            row["run_id"], row["dataset"], row["alpha"], row["min_require_size"],
            row["partition_seed"], row["training_seed"], row["status"], row["round_count"],
            row["first_round"], row["final_round"], row["best_round"], row["best_acc_pct"],
            row["best_ece"], row["best_nll"], row["final_acc_pct"], row["final_ece"],
            row["final_nll"], row["rounds96_100_acc_mean_pct"],
            row["rounds96_100_acc_sd_pct"], row["mean_round_time_s"],
            row["total_round_time_s"], row["console_log"], row["checkpoint"],
        ])
    style_table(ws)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = INPUT_FONT
        row[6].fill = GOOD_FILL if row[6].value == "complete" else WARN_FILL
    for column in ("C", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U"):
        for cell in ws[column][1:]:
            cell.number_format = "0.0000"
    set_widths(ws, {
        "A": 48, "B": 16, "C": 10, "D": 16, "E": 14, "F": 14, "G": 12,
        "H": 10, "I": 11, "J": 11, "K": 11, "L": 13, "M": 12, "N": 12,
        "O": 13, "P": 12, "Q": 12, "R": 20, "S": 17, "T": 19, "U": 19,
        "V": 72, "W": 72,
    })


def add_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("3-Seed Summary")
    headers = [
        "Dataset", "Alpha", "Min Require Size", "Completed Runs",
        "Best Acc TS42", "Best Acc TS43", "Best Acc TS44", "Best Mean (%)",
        "Best Sample SD (%)", "Best Min (%)", "Best Max (%)", "Final Mean (%)",
        "Final Sample SD (%)", "R96–100 Mean (%)", "Mean Round Time (s)",
        "Total Time (h)", "Status",
    ]
    ws.append(headers)
    detail_last = 145
    for dataset in DATASETS:
        for floor in FLOORS:
            for alpha in ALPHAS:
                row = ws.max_row + 1
                ws.append([dataset, alpha, floor])
                criteria = (
                    f"'Run Details'!$B$2:$B${detail_last},$A{row},"
                    f"'Run Details'!$C$2:$C${detail_last},$B{row},"
                    f"'Run Details'!$D$2:$D${detail_last},$C{row}"
                )
                ws.cell(row, 4, (
                    f'=COUNTIFS({criteria},'
                    f"'Run Details'!$G$2:$G${detail_last},\"complete\")"
                ))
                for offset, seed in enumerate(SEEDS, start=5):
                    ws.cell(row, offset, (
                        f'=IF($D{row}<>3,\"\",SUMIFS(\'Run Details\'!$L$2:$L${detail_last},'
                        f"{criteria},'Run Details'!$F$2:$F${detail_last},{seed}))"
                    ))
                ws.cell(row, 8, f'=IF($D{row}=3,AVERAGE(E{row}:G{row}),\"\")')
                ws.cell(row, 9, f'=IF($D{row}=3,STDEV(E{row}:G{row}),\"\")')
                ws.cell(row, 10, f'=IF($D{row}=3,MIN(E{row}:G{row}),\"\")')
                ws.cell(row, 11, f'=IF($D{row}=3,MAX(E{row}:G{row}),\"\")')
                ws.cell(row, 12, (
                    f'=IF($D{row}=3,AVERAGEIFS(\'Run Details\'!$O$2:$O${detail_last},'
                    f"{criteria}),\"\")"
                ))
                ws.cell(row, 13, (
                    f'=IF($D{row}=3,STDEV('
                    f"SUMIFS('Run Details'!$O$2:$O${detail_last},{criteria},'Run Details'!$F$2:$F${detail_last},42),"
                    f"SUMIFS('Run Details'!$O$2:$O${detail_last},{criteria},'Run Details'!$F$2:$F${detail_last},43),"
                    f"SUMIFS('Run Details'!$O$2:$O${detail_last},{criteria},'Run Details'!$F$2:$F${detail_last},44)),\"\")"
                ))
                ws.cell(row, 14, (
                    f'=IF($D{row}=3,AVERAGEIFS(\'Run Details\'!$R$2:$R${detail_last},'
                    f"{criteria}),\"\")"
                ))
                ws.cell(row, 15, (
                    f'=IF($D{row}=3,AVERAGEIFS(\'Run Details\'!$T$2:$T${detail_last},'
                    f"{criteria}),\"\")"
                ))
                ws.cell(row, 16, (
                    f'=SUMIFS(\'Run Details\'!$U$2:$U${detail_last},{criteria})/3600'
                ))
                ws.cell(row, 17, f'=IF(D{row}=3,\"Complete\",\"Incomplete\")')
    style_table(ws)
    for row in ws.iter_rows(min_row=2):
        for cell in row[:3]:
            cell.font = INPUT_FONT
        for cell in row[3:]:
            cell.font = FORMULA_FONT
        row[16].fill = GOOD_FILL if row[16].value == '=IF(D2=3,"Complete","Incomplete")' else PatternFill()
    for column in range(2, 17):
        for cell in ws[get_column_letter(column)][1:]:
            cell.number_format = "0.0000"
    ws["H1"].comment = Comment(
        "Mean of best global test accuracy across training seeds 42, 43, and 44.",
        "Codex",
    )
    ws["I1"].comment = Comment("Sample SD across the three training seeds.", "Codex")
    set_widths(ws, {
        "A": 16, "B": 10, "C": 17, "D": 16, "E": 14, "F": 14, "G": 14,
        "H": 15, "I": 20, "J": 14, "K": 14, "L": 15, "M": 20,
        "N": 18, "O": 20, "P": 15, "Q": 12,
    })


def add_floor_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Floor Pairing")
    headers = [
        "Dataset", "Alpha", "Training Seed", "Floor 0 Best Acc (%)",
        "Floor 1 Best Acc (%)", "Floor 1 − Floor 0 (pp)",
        "Floor 0 R96–100 (%)", "Floor 1 R96–100 (%)", "R96–100 Delta (pp)",
    ]
    ws.append(headers)
    detail_last = 145
    for dataset in DATASETS:
        for alpha in ALPHAS:
            for seed in SEEDS:
                row = ws.max_row + 1
                ws.append([dataset, alpha, seed])
                for floor, column in ((0, 4), (1, 5)):
                    ws.cell(row, column, (
                        f'=SUMIFS(\'Run Details\'!$L$2:$L${detail_last},'
                        f"'Run Details'!$B$2:$B${detail_last},$A{row},"
                        f"'Run Details'!$C$2:$C${detail_last},$B{row},"
                        f"'Run Details'!$D$2:$D${detail_last},{floor},"
                        f"'Run Details'!$F$2:$F${detail_last},$C{row})"
                    ))
                ws.cell(row, 6, f'=E{row}-D{row}')
                for floor, column in ((0, 7), (1, 8)):
                    ws.cell(row, column, (
                        f'=SUMIFS(\'Run Details\'!$R$2:$R${detail_last},'
                        f"'Run Details'!$B$2:$B${detail_last},$A{row},"
                        f"'Run Details'!$C$2:$C${detail_last},$B{row},"
                        f"'Run Details'!$D$2:$D${detail_last},{floor},"
                        f"'Run Details'!$F$2:$F${detail_last},$C{row})"
                    ))
                ws.cell(row, 9, f'=H{row}-G{row}')
    style_table(ws)
    for row in ws.iter_rows(min_row=2):
        for cell in row[:3]:
            cell.font = INPUT_FONT
        for cell in row[3:]:
            cell.font = FORMULA_FONT
    for column in "BDEFGHI":
        for cell in ws[column][1:]:
            cell.number_format = "0.0000"
    set_widths(ws, {"A": 16, "B": 10, "C": 15, "D": 21, "E": 21, "F": 24, "G": 21, "H": 21, "I": 21})


def add_trend_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Trend Charts")
    ws.sheet_view.showGridLines = False
    ws.append(["Dataset", "Alpha", "Floor 0 Best Mean (%)", "Floor 1 Best Mean (%)"])
    summary_last = 49
    for dataset in DATASETS:
        for alpha in ALPHAS:
            row = ws.max_row + 1
            ws.append([dataset, alpha])
            for floor, column in ((0, 3), (1, 4)):
                ws.cell(row, column, (
                    f'=SUMIFS(\'3-Seed Summary\'!$H$2:$H${summary_last},'
                    f"'3-Seed Summary'!$A$2:$A${summary_last},$A{row},"
                    f"'3-Seed Summary'!$B$2:$B${summary_last},$B{row},"
                    f"'3-Seed Summary'!$C$2:$C${summary_last},{floor})"
                ))
    style_table(ws)
    for row in ws.iter_rows(min_row=2):
        row[0].font = INPUT_FONT
        row[1].font = INPUT_FONT
        row[2].font = FORMULA_FONT
        row[3].font = FORMULA_FONT
    for cell in ws["B"][1:] + ws["C"][1:] + ws["D"][1:]:
        cell.number_format = "0.0000"
    set_widths(ws, {"A": 16, "B": 10, "C": 24, "D": 24})

    for dataset_index, dataset in enumerate(DATASETS):
        start = 2 + dataset_index * len(ALPHAS)
        end = start + len(ALPHAS) - 1
        chart = LineChart()
        chart.title = f"{dataset}: Best Accuracy vs Alpha"
        chart.y_axis.title = "Accuracy (%)"
        chart.x_axis.title = "Dirichlet alpha"
        chart.style = 13
        chart.height = 8
        chart.width = 15
        chart.add_data(
            Reference(ws, min_col=3, max_col=4, min_row=start, max_row=end),
            titles_from_data=False,
        )
        chart.series[0].tx = SeriesLabel(v="Floor 0 Best Mean (%)")
        chart.series[1].tx = SeriesLabel(v="Floor 1 Best Mean (%)")
        chart.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
        ws.add_chart(chart, f"F{2 + dataset_index * 16}")


def add_audit_sheet(wb: Workbook, audit_rows: list[tuple[str, object, str]]) -> None:
    ws = wb.create_sheet("Audit")
    ws.append(["Check", "Observed", "Result"])
    for check, observed, result in audit_rows:
        ws.append([check, observed, result])
    style_table(ws)
    for row in ws.iter_rows(min_row=2):
        row[0].font = INPUT_FONT
        row[1].font = INPUT_FONT
        row[2].font = INPUT_FONT
        row[2].fill = GOOD_FILL if row[2].value == "PASS" else WARN_FILL
    set_widths(ws, {"A": 52, "B": 72, "C": 14})


def build_workbook(path: Path, runs: list[dict[str, object]], audit_rows: list[tuple[str, object, str]]) -> None:
    wb = Workbook()
    add_readme_sheet(wb, runs)
    add_run_sheet(wb, runs)
    add_summary_sheet(wb)
    add_floor_sheet(wb)
    add_trend_sheet(wb)
    add_audit_sheet(wb, audit_rows)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.font.name != "Arial":
                    cell.font = Font(name="Arial", size=10, color=cell.font.color)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(path)


def write_report(path: Path, runs: list[dict[str, object]], summary: list[dict[str, object]], pairs: list[dict[str, object]]) -> None:
    completed = sum(row["status"] == "complete" for row in runs)
    lines = [
        "# FedAvg ViT-B/16 144-Run Summary",
        "",
        f"- Completed runs: {completed}/{len(runs)}.",
        "- Primary metric: best global test accuracy within 100 rounds.",
        "- Values are mean ± sample SD across training seeds 42, 43, and 44.",
        "- Partition seed is fixed at 42.",
        "",
        "## Dataset-level descriptive summary",
        "",
        "| Dataset | Floor | Mean across 8 configuration means (%) | Best configuration | Worst configuration |",
        "|---|---:|---:|---|---|",
    ]
    for dataset in DATASETS:
        for floor in FLOORS:
            group = [row for row in summary if row["dataset"] == dataset and row["min_require_size"] == floor]
            values = [float(row["best_acc_mean_pct"]) for row in group]
            best = max(group, key=lambda row: float(row["best_acc_mean_pct"]))
            worst = min(group, key=lambda row: float(row["best_acc_mean_pct"]))
            lines.append(
                f"| {dataset} | {floor} | {mean(values):.2f} | "
                f"alpha={float(best['alpha']):.2f}, {float(best['best_acc_mean_pct']):.2f}% | "
                f"alpha={float(worst['alpha']):.2f}, {float(worst['best_acc_mean_pct']):.2f}% |"
            )
    lines.extend([
        "",
        "## Floor sensitivity",
        "",
        "Positive values mean `min_require_size=1` is higher than `min_require_size=0`.",
        "",
        "| Dataset | Mean paired delta (pp) | Median paired delta (pp) | Range (pp) |",
        "|---|---:|---:|---:|",
    ])
    for dataset in DATASETS:
        values = [float(row["floor1_minus_floor0_best_pp"]) for row in pairs if row["dataset"] == dataset]
        lines.append(
            f"| {dataset} | {mean(values):.2f} | {median(values):.2f} | "
            f"{min(values):.2f} to {max(values):.2f} |"
        )
    seed_sds = [float(row["best_acc_sample_sd_pct"]) for row in summary]
    lines.extend([
        "",
        "## Reproducibility and limitations",
        "",
        f"- Median three-seed sample SD across 48 configurations: {median(seed_sds):.3f} percentage points.",
        f"- Maximum three-seed sample SD: {max(seed_sds):.3f} percentage points.",
        "- Alpha and floor comparisons are descriptive because only one partition seed is used.",
        "- The workbook retains best-round, Round-100, and rounds-96–100 metrics so downstream reporting can use an explicit metric definition.",
        "",
    ])
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    matrix = expected_matrix()
    runs = [parse_run(row) for row in matrix]
    summary = aggregate_runs(runs)
    pairs = build_floor_pairs(runs) if all(row["status"] == "complete" for row in runs) else []

    observed_ids = [str(row["run_id"]) for row in runs]
    expected_ids = [row[0] for row in matrix]
    audit_rows = [
        ("Expected matrix size", len(expected_ids), "PASS" if len(expected_ids) == 144 else "FAIL"),
        ("Unique expected run IDs", len(set(expected_ids)), "PASS" if len(set(expected_ids)) == 144 else "FAIL"),
        ("Parsed run rows", len(runs), "PASS" if len(runs) == 144 else "FAIL"),
        ("Duplicate parsed run IDs", sum(count - 1 for count in Counter(observed_ids).values() if count > 1), "PASS" if len(set(observed_ids)) == 144 else "FAIL"),
        ("Strictly complete runs", sum(row["status"] == "complete" for row in runs), "PASS" if all(row["status"] == "complete" for row in runs) else "FAIL"),
        ("Runs with exactly rounds 1–100", sum(row["round_count"] == 100 and row["unique_round_count"] == 100 and row["first_round"] == 1 and row["final_round"] == 100 for row in runs), "PASS" if all(row["round_count"] == 100 and row["unique_round_count"] == 100 and row["first_round"] == 1 and row["final_round"] == 100 for row in runs) else "FAIL"),
        ("Best summary agrees with round log", sum(bool(row["best_matches_round_log"]) for row in runs), "PASS" if all(row["best_matches_round_log"] for row in runs) else "FAIL"),
        ("Configuration metadata matches matrix", sum(bool(row["config_ok"]) for row in runs), "PASS" if all(row["config_ok"] for row in runs) else "FAIL"),
        ("Checkpoint files present", sum(bool(row["checkpoint_exists"]) for row in runs), "PASS" if all(row["checkpoint_exists"] for row in runs) else "FAIL"),
        ("Three completed seeds per aggregate", min(int(row["n_complete"]) for row in summary), "PASS" if all(row["n_complete"] == 3 for row in summary) else "FAIL"),
    ]

    write_csv(OUTPUT_ROOT / "fedavg_vit_runs.csv", runs, RUN_FIELDS)
    write_csv(OUTPUT_ROOT / "fedavg_vit_summary.csv", summary, SUMMARY_FIELDS)
    write_csv(OUTPUT_ROOT / "fedavg_vit_floor_pairs.csv", pairs, PAIR_FIELDS)
    write_report(OUTPUT_ROOT / "report.md", runs, summary, pairs)
    build_workbook(OUTPUT_ROOT / "FedAvg_ViT_alpha_144_results.xlsx", runs, audit_rows)

    audit_payload = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "results_root": str(RESULTS_ROOT.relative_to(ROOT)),
        "checks": [
            {"check": check, "observed": observed, "result": result}
            for check, observed, result in audit_rows
        ],
    }
    (OUTPUT_ROOT / "audit.json").write_text(
        json.dumps(audit_payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    failed = [check for check, _, result in audit_rows if result != "PASS"]
    if failed:
        raise SystemExit("Audit failed: " + ", ".join(failed))
    print(f"Wrote FedAvg summary package to {OUTPUT_ROOT}")


if __name__ == "__main__":
    main()
